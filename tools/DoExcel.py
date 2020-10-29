from openpyxl import load_workbook
from tools.FilePath import *
from tools.read_conf import ReadConf
from tools.random_chinese import GBK2312
import json
from tools.mysql import MysqlCon



# 创建读取/写入测试数据类
class ReadTestData():

    # 创建读取测试数据方法
    def read_test_data(self, filename):
        # 打开文件
        wb = load_workbook(filename)
        name = str(filename)
        if "tray" in name:
            mode = eval(ReadConf().read_conf(conf_path, "MODE", "tray_mode"))
        elif "logistics" in name:
            mode = eval(ReadConf().read_conf(conf_path, "MODE", "logistice_mode"))
        elif "warehouse" in name:
            mode = eval(ReadConf().read_conf(conf_path, "MODE", "warehouse_mode"))


        # 定义列表接收测试数据
        test_data = []
        for key in mode:
            # 每个key都是一个表名
            sheet = wb[key]  # 打开sheet表
            # 循环取值
            for i in range(2, sheet.max_row+1):
                # 定义字典，接收每一行的内容
                sub_data = {}
                sub_data["case_id"] = sheet.cell(i, 1).value
                sub_data["title"] = sheet.cell(i, 2).value
                sub_data["pri"] = sheet.cell(i, 3).value
                sub_data["url"] = sheet.cell(i, 4).value
                sub_data['data'] = sheet.cell(i, 5).value
                sub_data['method'] = sheet.cell(i, 6).value
                sub_data['expected'] = sheet.cell(i, 7).value
                sub_data['need_random_telephone'] = sheet.cell(i, 8).value
                sub_data['need_random_companyname'] = sheet.cell(i, 9).value
                sub_data['rely_on_case_id'] = sheet.cell(i, 10).value
                sub_data['rely_on_value'] = sheet.cell(i, 11).value
                sub_data['result'] = sheet.cell(i, 12).value
                sub_data['sheet_name'] = key
                if mode[key] == 'all':  # 该sheet表中需要执行所有用例
                    test_data.append(sub_data)
                elif sub_data["pri"] in mode[key]:
                    test_data.append(sub_data)
        return test_data


    # 创建写入数据方法，将测试结果写入excel
    def write_result(self, filename, sheet_name, i, value, test_result):
        # 打开文件
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        # 写入数据(只能写入str数据类型)
        sheet.cell(i, 12).value = str(value)
        sheet.cell(i, 13).value = str(test_result)
        # 保存文件
        wb.save(filename)

    # 创建获取指定行、指定列的值方法
    def get_value(self, filename, sheet_name, row, column):
        # 打开文件
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        # 读数据
        value = sheet.cell(row, column).value
        return value

    # 随机获取一个公司名，判断是否重复，写入excel数据中
    def write_customer_name(self, filename):
        wb = load_workbook(filename)
        sheet = wb['admin_add_customer_C']
        for j in range(2, sheet.max_row+1):
            first_name = eval(self.get_value(filename, 'admin_add_customer_C', j, 5))
            first_name['companyName'] = '上海市自动化测试公司{0}'.format(GBK2312())
            # 查询数据库，根据公司名称判断该公司是否已存在，若存在，则重新获取随机值
            select_result = MysqlCon().select('SELECT * FROM t_customer_company WHERE chinese_name="%s"'%first_name['companyName'],'test_yelo_basicdata')
            len1 = len(select_result)
            if len1 > 0: # 说明已存在相应的数据
                self.write_customer_name(filename)
            else:
                sheet.cell(j, 5).value = json.dumps(first_name, ensure_ascii=False)
                wb.save(filename)
        # return sheet.max_row

    def write_deff_telephone(self,filename):
        test_data = self.read_test_data(filename)
        wb = load_workbook(filename)
        for data in test_data:
            if data['need_random_telephone']:
                sheet = wb[data['sheet_name']]
                read_data = json.loads(data['data'])
                # 获取11位的手机号
                random_mobile = '1'
                for k in range(10):
                    random_mobile += str(k)
                # 通过数据库查询随机生成的手机号是否已注册

                select_result_tele = MysqlCon().select("SELECT ID, AES_DECRYPT(UNHEX(user_account),'1W5hoD!qU^5&L#Ix') FROM t_account WHERE AES_DECRYPT(UNHEX(user_account),'1W5hoD!qU^5&L#Ix')='%s'"%random_mobile,'test_yelo_basicdata')
                if select_result_tele:
                    self.write_deff_telephone(filename)
                else:
                    read_data['mobile'] = random_mobile
                    wb.save(filename)
                    return read_data


    def read_result(self, filename, sheet_name, case_id,rely_on_value):
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        test_data = self.read_test_data(filename)
        #定义一个空列表接收对应sheet的数据
        sheet_test_data = []
        for sheet_data in test_data:
            if sheet_data['sheet_name'] == sheet_name:
                sheet_test_data.append(sheet_data)
        #将rely_on_value存在一个列表中
        rely_on_value = str(rely_on_value)
        rely_on_value_list = rely_on_value.split(',')
        #定义空字典接收数据
        result_dict = {}
        for data in sheet_test_data:
            if data['case_id'] == case_id:
                result = eval(data['result'])
                for k in rely_on_value_list:
                    result_dict[k] = result['data'][k]
        #返回为字典格式
        return result_dict








if __name__ == '__main__':
    # re1 = ReadTestData().read_test_data('C:\\Users\\17826\\PycharmProjects\\2020_api_auto\\test_data\\test_tray.xlsx')
    # re2 = ReadTestData().write_customer_name('C:\\Users\\17826\\PycharmProjects\\2020_api_auto\\test_data\\test_tray.xlsx')
    re3 = ReadTestData().write_deff_telephone('C:\\Users\\17826\\PycharmProjects\\2020_api_auto\\test_data\\test_tray.xlsx')
    re4 = ReadTestData().read_result('C:\\Users\\17826\\PycharmProjects\\2020_api_auto\\test_data\\test_tray.xlsx','web_sign_up_in',1,'random,verificationId')
    print(re4)












